"""
Excel reader — multi-sheet, fully dynamic.
Master sheet: client static info keyed by ic_number.
Batch sheets: transaction-specific data joined to Master by ic_number.

P1-1: All load_workbook() calls protected against PermissionError
       (Excel holds exclusive lock on Windows when file is open).
Fix-8: IC numbers normalized both sides of join.
Fix-9: Native Python types preserved through reader.
"""
import re
import openpyxl
from pathlib import Path


# ── IC normalization ──────────────────────────────────────────────────────────

def normalize_ic(raw) -> str:
    """Strip non-digits, zero-pad to 12. Empty → ''.
    Handles float from Excel (880101145678.0 → '880101145678').
    """
    # Convert float to int string first to avoid '880101145678.0'
    if isinstance(raw, float):
        raw = int(raw)
    cleaned = re.sub(r"[^0-9]", "", str(raw))
    return cleaned.zfill(12) if cleaned else ""


# ── Safe workbook open ────────────────────────────────────────────────────────

class ExcelLockedError(Exception):
    """Raised when Excel holds an exclusive lock on the file."""


def _open_wb(path: Path):
    """Open workbook, converting Windows PermissionError to ExcelLockedError."""
    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=True)
    except PermissionError:
        raise ExcelLockedError(
            f"Cannot open '{path.name}' — it is locked by Excel.\n"
            "Close the file in Excel, then click ↻ Reload."
        )


# ── Sheet parsing ─────────────────────────────────────────────────────────────

def _sheet_to_dicts(ws) -> list[dict]:
    """
    Parse a worksheet. Native types (datetime, int, float) preserved.
    None cells → "". Strings stripped. ic_number auto-normalized.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [
        str(h).strip().lower().replace(" ", "_") if h else f"col_{i}"
        for i, h in enumerate(rows[0])
    ]
    result = []
    for row in rows[1:]:
        if not any(row):
            continue
        record = {}
        for k, v in zip(headers, row):
            if v is None:
                record[k] = ""
            elif isinstance(v, str):
                record[k] = v.strip()
            else:
                record[k] = v  # int, float, datetime preserved for formatters
        if "ic_number" in record:
            record["ic_number"] = normalize_ic(record["ic_number"])
        result.append(record)
    return result


# ── Public loaders ────────────────────────────────────────────────────────────

def load_master(path: Path) -> dict[str, dict]:
    """Returns {normalized_ic: client_dict} from Master sheet."""
    wb = _open_wb(path)
    clients = _sheet_to_dicts(wb["Master"])
    wb.close()
    return {c["ic_number"]: c for c in clients if c.get("ic_number")}


def get_sheet_headers(path: Path, sheet_name: str) -> list[str]:
    """Column names from a batch sheet, excluding ic_number."""
    wb = _open_wb(path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
    wb.close()
    return [
        str(h).strip().lower().replace(" ", "_")
        for h in first_row
        if h and str(h).strip().lower().replace(" ", "_") != "ic_number"
    ]


def load_batch(path: Path, sheet_name: str,
               master: dict[str, dict]) -> list[dict]:
    """Master + batch data merged by ic_number. Missing Master rows skipped."""
    wb = _open_wb(path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
        )
    rows = _sheet_to_dicts(wb[sheet_name])
    wb.close()
    merged = []
    for row in rows:
        ic = row.get("ic_number", "")
        master_rec = master.get(ic)
        if not master_rec:
            continue
        merged.append({**master_rec, **row})
    return merged


def find_client_in_batch(path: Path, sheet_name: str, ic_number: str) -> dict:
    """Return batch row for one client (normalized IC), or {} if not found."""
    ic_norm = normalize_ic(ic_number)
    try:
        wb = _open_wb(path)
    except ExcelLockedError:
        return {}
    if sheet_name not in wb.sheetnames:
        wb.close()
        return {}
    rows = _sheet_to_dicts(wb[sheet_name])
    wb.close()
    for row in rows:
        if row.get("ic_number") == ic_norm:
            return row
    return {}


def get_master_names(master: dict[str, dict]) -> list[str]:
    return sorted(c.get("name", "") for c in master.values() if c.get("name"))


def sheet_names(path: Path) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names
